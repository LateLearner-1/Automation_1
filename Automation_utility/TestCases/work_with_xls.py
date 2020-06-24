import openpyxl
# read file:
# load workbook
# wk=openpyxl.load_workbook('C:\\Users\\HP\\Desktop\\Training status\\Asset_Allocating.xlsx')
#
# print('number of sheets: '+str(wk.sheetnames))
# print('Active sheet as object: ' + str(wk.active))
# print('Active sheet as object name: ' + str(wk.active.title))
# # Create object of sheet on which you want to work
#
# obj=wk['Sheet1']
# # #To fetch data from cell
# # print(obj['A5'].value)
# # # Another way to fetch value is as below
# # print(obj.cell(5,1).value)
# # # look at the value as input in above two commands and keep in mind, following is one more method
# # print(obj.cell(column=1, row=5).value)
# #
# # #now read all data from sheet
# #
# # #Step 1: find rows and columns having data
# # rows = obj.max_row
# # columns = obj.max_column
# # print(rows, columns)
# # d=[]
# # # Code to print each row as sparate comma separated row
# # for i in range(1, rows+1):
# #     c=[]
# #     for j in range(1, columns+1):
# #         c.append(obj.cell(i, j).value)
# #     d.append(c)
# # print(d)

# Write data to excel
wk2 = openpyxl.Workbook()
obj1= wk2.active
obj1.title="to_be_saved"
obj1['B5'].value='first_Entry'

# Second sheet created
wk2.create_sheet('to_sheet2')
obj2 = wk2['to_sheet2']
obj2['B6'].value='second entry'

#Remove sheet
wk2.remove(wk2['to_be_saved'])

#Save to workbook
wk2.save('C:\\Users\\HP\\Desktop\\Training status\\Sheet_to_write.xlsx')
wk2.create_sheet('to_sheet2')
