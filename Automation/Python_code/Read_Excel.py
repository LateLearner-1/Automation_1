import openpyxl

wk = openpyxl.load_workbook('C:\\Users\\HP\\PycharmProjects\\Automation\\Python_code\\TestData_1.xlsx')

def fetch_number_of_rows(sheet_name):
    sheet = wk[sheet_name]
    return sheet.max_row

def fetch_cell_data(sheet_name, row, cell):
    sheet = wk[sheet_name]
    data = sheet.cell(int(row), int(cell))
    return data.value

a = fetch_cell_data('Sheet1', 2, 3)
print(a)
